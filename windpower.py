# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'main.ui'
#
# Created by: PyQt5 UI code generator 5.15.2
#
# WARNING: Any manual changes made to this file will be lost when pyuic5 is
# run again.  Do not edit this file unless you know what you are doing.

from PyQt5 import QtCore, QtGui, QtWidgets
import sys
import qtmodern.styles
import qtmodern.windows
import pyowm
import openpyxl
from PyQt5.QtWidgets import QMessageBox, QWidget, QToolTip, QPushButton, QApplication
from PyQt5.QtGui import QFont
from forall import Ui_Form

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(646, 520)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("main.png"), QtGui.QIcon.Normal, QtGui.QIcon.On)
        MainWindow.setWindowIcon(icon)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(150, 30, 391, 71))
        font = QtGui.QFont()
        font.setPointSize(28)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(40, 120, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(40, 190, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(40, 330, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.comboBox = QtWidgets.QComboBox(self.centralwidget)
        self.comboBox.setGeometry(QtCore.QRect(170, 120, 371, 41))
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItem("---SELECT THE REGION---")
        self.comboBox.addItem("---FOR ALL REGIONS---")
        self.comboBox.addItem("Akmola Region")
        self.comboBox.addItem("Aktobe Region")
        self.comboBox.addItem("Almaty")
        self.comboBox.addItem("Almaty Region")
        self.comboBox.addItem("Atyrau Region")
        self.comboBox.addItem("East Kazakhstan Region")
        self.comboBox.addItem("Jambyl Region")
        self.comboBox.addItem("Karagandy Region")
        self.comboBox.addItem("Kostanay Region")
        self.comboBox.addItem("Kyzylorda Region")
        self.comboBox.addItem("Mangystau Region")
        self.comboBox.addItem("North Kazakhstan Region")
        self.comboBox.addItem("Nur-Sultan")
        self.comboBox.addItem("Pavlodar Region")
        self.comboBox.addItem("Shymkent")
        self.comboBox.addItem("Turkistan Region")
        self.comboBox.addItem("West Kazakhstan Region")

        self.areaLabel = QtWidgets.QLineEdit(self.centralwidget)
        self.areaLabel.setGeometry(QtCore.QRect(170, 190, 371, 41))
        self.areaLabel.setAcceptDrops(True)
        self.areaLabel.setAutoFillBackground(False)
        self.areaLabel.setInputMethodHints(QtCore.Qt.ImhPreferNumbers)
        self.areaLabel.setInputMask("")
        self.areaLabel.setReadOnly(True)
        self.areaLabel.setPlaceholderText("")
        self.areaLabel.setStyleSheet("color: grey;")
        self.areaLabel.setClearButtonEnabled(True)
        self.areaLabel.setObjectName("areaLabel")
        self.lineEdit_2 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_2.setGeometry(QtCore.QRect(170, 330, 371, 41))
        self.lineEdit_2.setInputMethodHints(QtCore.Qt.ImhPreferNumbers)
        self.lineEdit_2.setClearButtonEnabled(True)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(230, 420, 181, 41))
        self.pushButton.setObjectName("pushButton")
        self.label_5 = QtWidgets.QLabel(self.centralwidget)
        self.label_5.setGeometry(QtCore.QRect(550, 190, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(self.centralwidget)
        self.label_6.setGeometry(QtCore.QRect(550, 330, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.label_7 = QtWidgets.QLabel(self.centralwidget)
        self.label_7.setGeometry(QtCore.QRect(550, 260, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_7.setFont(font)
        self.label_7.setObjectName("label_7")
        self.label_8 = QtWidgets.QLabel(self.centralwidget)
        self.label_8.setGeometry(QtCore.QRect(40, 260, 141, 41))
        font = QtGui.QFont()
        font.setPointSize(16)
        self.label_8.setFont(font)
        self.label_8.setObjectName("label_8")
        self.areaLabel_2 = QtWidgets.QLineEdit(self.centralwidget)
        self.areaLabel_2.setGeometry(QtCore.QRect(170, 260, 371, 41))
        self.areaLabel_2.setAcceptDrops(True)
        self.areaLabel_2.setAutoFillBackground(False)
        self.areaLabel_2.setInputMethodHints(QtCore.Qt.ImhPreferNumbers)
        self.areaLabel_2.setInputMask("")
        self.areaLabel_2.setText("")
        self.areaLabel_2.setPlaceholderText("")
        self.areaLabel_2.setClearButtonEnabled(True)
        self.areaLabel_2.setObjectName("areaLabel_2")
        self.pushButton_3 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_3.setGeometry(QtCore.QRect(580, 20, 41, 41))
        self.pushButton_3.setText("")
        self.pushButton_3.setToolTip('About the program')
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap("about.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_3.setIcon(icon1)
        self.pushButton_3.setIconSize(QtCore.QSize(25, 25))
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_4 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_4.setGeometry(QtCore.QRect(580, 70, 41, 41))
        self.pushButton_4.setText("")
        self.pushButton_4.setToolTip('History of Calculations')
        icon2 = QtGui.QIcon()
        icon2.addPixmap(QtGui.QPixmap("history.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_4.setIcon(icon2)
        self.pushButton_4.setIconSize(QtCore.QSize(25, 25))
        self.pushButton_4.setObjectName("pushButton_4")
        self.pushButton_5 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_5.setGeometry(QtCore.QRect(580, 120, 41, 41))
        self.pushButton_5.setText("")
        self.pushButton_5.setToolTip('Calculate for All Regions')
        icon3 = QtGui.QIcon()
        icon3.addPixmap(QtGui.QPixmap("all.jpg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_5.setIcon(icon3)
        self.pushButton_5.setIconSize(QtCore.QSize(25, 25))
        self.pushButton_5.setObjectName("pushButton_5")
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.actionOne_Region = QtWidgets.QAction(MainWindow)
        self.actionOne_Region.setObjectName("actionOne_Region")
        self.actionAll_Regions = QtWidgets.QAction(MainWindow)
        self.actionAll_Regions.setObjectName("actionAll_Regions")
        self.actionLast_update = QtWidgets.QAction(MainWindow)
        self.actionLast_update.setObjectName("actionLast_update")

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Wind Power Calculator"))
        self.label.setText(_translate("MainWindow", "Power Calculation"))
        self.label_2.setText(_translate("MainWindow", "Region"))
        self.label_3.setText(_translate("MainWindow", "Speed"))
        self.label_4.setText(_translate("MainWindow", "Air density"))
        self.areaLabel.setText(_translate("MainWindow", "No need to fill: data taken online"))
        self.lineEdit_2.setText(_translate("MainWindow", "1.24"))
        self.pushButton.setText(_translate("MainWindow", "Calculate"))
        self.pushButton.clicked.connect(self.calculate)
        #self.pushButton_5.clicked.connect(self.openForm)
        self.label_5.setText(_translate("MainWindow", "m/s"))
        self.label_6.setText(_translate("MainWindow", "kg/m3"))
        self.label_7.setText(_translate("MainWindow", "m2"))
        self.label_8.setText(_translate("MainWindow", "Area"))

    def calculate(self):
        owm = pyowm.OWM('a4059161b180b05d8caa7275375bdf1d')
        mgr = owm.weather_manager()
        region = str(self.comboBox.currentText())
        area = self.areaLabel_2.text()
        density = self.lineEdit_2.text()
        try:
            number = int(area)
            city = str(self.reg_select(region))
            if((city!="default") and (city!="all")):
                loc = mgr.weather_at_place(city)
                weather = loc.weather
                wind = weather.wind()
                speed = wind.get('speed')
                self.areaLabel.setText(str(speed))
                power = (0.5*float(density)*float(area)*(speed**3))/1000
                self.label.setText("Power is: "+str(round(power, 2))+" kW")
                self.update()
                filename = 'base.xlsx'
                wb = openpyxl.load_workbook(filename)
                ws = wb.active
                last_row = ws.max_row+1
                ws.cell(column=1, row=last_row, value=region)
                ws.cell(column=2, row=last_row, value=speed)
                ws.cell(column=3, row=last_row, value=area)
                ws.cell(column=4, row=last_row, value=density)
                ws.cell(column=5, row=last_row, value=power)
                wb.save(filename)
            elif(city=="default"):
                d = QMessageBox()
                d.setWindowTitle("Region is not selected")
                d.setText("Please, select the region from the dropdown menu")
                d.setIcon(QMessageBox.Warning)
                x = d.exec_()
                pass
            elif(city=="all"):
                print("CHECK 1")
                cities = ["Kokshetau", "Aktobe", "Almaty", "Taldykorgan", "Atyrau", "Oskemen", "Taraz", "Karagandy", "Kostanay", "Kyzylorda", "Aktau", "Petropavl", "Nur-Sultan", "Pavlodar", "Shymkent", "Turkistan", "Oral"]
                self.total = []
                for each in cities:
                    print(each)
                    loc = mgr.weather_at_place(each)
                    weather = loc.weather
                    wind = weather.wind()
                    speed = wind.get('speed')
                    #self.areaLabel.setText(str(speed))
                    power = (0.5*float(density)*float(area)*(speed**3))/1000
                    self.total.append(str(power))

                self.openForm()
                #for each in cities:

        except Exception:
            msg = QMessageBox()
            msg.setWindowTitle("Wrong Input")
            msg.setText("Invalid Input type. Please, enter the values properly.")
            msg.setIcon(QMessageBox.Warning)
            msg.setDetailedText("- Decimal point has to be seperated by the dot(.), not comma (,). \n - Area and Density boxes cannot be empty.")
            x = msg.exec_()
            pass
        #if(1):
        #    cities = ["Aktobe", "Kokshetau", "Almaty", "Taldykorgan", "Atyrau", "Oskemen", "Taraz", "Karagandy", "Kostanay", "Kyzylorda", "Aktau", "Petropavl", "Nur-Sultan", "Pavlodar", "Shymkent", "Turkistan", "Oral"]
        #    self.openForm(cities)
    def update(self):
        self.label.adjustSize()

    def openForm(self):
        self.window = QtWidgets.QMainWindow()
        self.message = self.total
        self.ui = Ui_Form(self.message)
        self.ui.setupUi(self.window)
        self.window.show()

    def reg_select(self, reg):
        if(reg=="Aktobe Region"):
            return "Aktobe"
        elif(reg=="---SELECT THE REGION---"):
            return "default"
        elif(reg=="---FOR ALL REGIONS---"):
            return "all"
        elif(reg=="Akmola Region"):
            return "Kokshetau"
        elif(reg=="Almaty"):
            return "Almaty"
        elif(reg=="Almaty Region"):
            select = "Taldykorgan"
        elif(reg=="Atyrau Region"):
            select = "Atyrau"
        elif(reg=="East Kazakhstan Region"):
            select = "Oskemen"
        elif(reg=="Jambyl Region"):
            select = "Taraz"
        elif(reg=="Karagandy Region"):
            select = "Karagandy"
        elif(reg=="Kostanay Region"):
            select = "Kostanay"
        elif(reg=="Kyzylorda Region"):
            select = "Kyzylorda"
        elif(reg=="Mangystau Region"):
            select = "Aktau"
        elif(reg=="North Kazakhstan Region"):
            return "Petropavl"
        elif(reg=="Nur-Sultan"):
            return "Nur-Sultan"
        elif(reg=="Pavlodar Region"):
            return "Pavlodar"
        elif(reg=="Shymkent"):
            return "Shymkent"
        elif(reg=="Turkistan Region"):
            return "Turkistan"
        elif(reg=="West Kazakhstan Region"):
            return "Oral"

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
